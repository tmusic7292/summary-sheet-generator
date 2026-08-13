import os
os.environ["NUMBA_DISABLE_JIT"] = "1"

#!/usr/bin/env python3
"""
Summary Sheet Tool v13.0 — Fixed & Enhanced
Features:
- Real-time extraction via watchdog
- Remark = filename
- No skipping; all files processed
- NEW: S_NOx and S_THC columns from Specific Results
- NEW: Proper CO extraction from Specific Results
- All original features preserved
- Modern responsive dark UI
"""

import os
import re
import time
import threading
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pdfplumber
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# --- CONFIG ---
APP_TITLE = "Summary Sheet Tool v13.0 — Fixed Edition"
ADDITIONAL_HEADERS = [
    "Injector 1", "Injector 2", "Injector 3", "Injector 4",
    "DOC", "DPF", "Engine Number", "Dataset"
]
# Original headers PLUS new S_NOx and S_THC
HEADERS = ["S No.", "Date", "Test ID", "Test Name", "NOx", "THC", "CO", "PM", "PN", "S_NOx", "S_THC", "S_CO", "Remark"] + ADDITIONAL_HEADERS

# --- COLOR THEME (Modern Dark) ---
THEME = {
    "bg_dark": "#1a1a1a",
    "bg_lighter": "#242424",
    "bg_input": "#2d2d2d",
    "fg_text": "#e8e8e8",
    "fg_accent": "#00d4ff",
    "fg_success": "#00ff88",
    "fg_error": "#ff4466",
    "btn_bg": "#0066cc",
    "btn_hover": "#0080ff",
}

# ========== ENHANCED PDF PARSING ==========
def parse_pdf_comprehensive(path: str):
    """
    Extract data from:
    1. Emission Limits (NOx, THC, CO, PM, PN) - Original extraction
    2. Specific Results - S_NOx, S_THC, S_CO (NEW)
    """
    # Original extraction
    nox = ""
    thc = ""
    co = ""
    pm = ""
    pn = ""
    remark = ""
    remark_bits = []
    
    # New specific results extraction
    s_nox = ""
    s_thc = ""
    s_co = ""
    
    try:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                txt = page.extract_text() or ""
                lines = txt.splitlines()
                
                # ===== SECTION 1: Emission Limits (Original) =====
                if "Emission limits - Stage V NRE" in txt:
                    for l in lines:
                        nums = re.findall(r"([0-9][0-9Ee\+\-\.]*)", l)
                        
                        if "CO uncorrected" in l or "CO " in l:
                            co = nums[0] if nums else ""
                        elif "NOX" in l or "NOx" in l:
                            nox = nums[0] if nums else ""
                        elif "Particulate mass" in l:
                            pm = nums[0] if nums else ""
                        elif "Particulate number" in l:
                            pn = nums[0] if nums else ""
                        elif "THC" in l and "NO2" not in l:
                            thc = nums[0] if nums else ""
                        
                        if ("Passed" in l) or ("Failed" in l):
                            remark_bits.append(l.strip())
                
                # ===== SECTION 2: Specific Results (NEW) =====
                # Look for "Specific results" section
                if "Specific results" in txt.lower():
                    for i, line in enumerate(lines):
                        # Extract S_NOx
                        if re.search(r'\bNOX\b.*\[g/kWh\]', line, re.IGNORECASE) or re.search(r'\bNOx\b.*\[g/kWh\]', line, re.IGNORECASE):
                            # Get value from this line
                            parts = line.split()
                            for j, part in enumerate(parts):
                                if re.match(r'^[\d.]+$', part):
                                    s_nox = part
                                    break
                        
                        # Extract S_THC
                        if re.search(r'\bTHC\b.*\[g/kWh\]', line, re.IGNORECASE):
                            parts = line.split()
                            for j, part in enumerate(parts):
                                if re.match(r'^[\d.]+$', part):
                                    s_thc = part
                                    break
                        
                        # Extract S_CO - from "CO (COL - Tailpipe)" or similar
                        if re.search(r'CO.*\(.*Tailpipe\)', line, re.IGNORECASE):
                            parts = line.split()
                            for j, part in enumerate(parts):
                                if re.match(r'^[\d.]+$', part):
                                    s_co = part
                                    break
                        elif re.search(r'\bCO\b.*\[g/kWh\]', line, re.IGNORECASE) and "COL" not in line:
                            parts = line.split()
                            for j, part in enumerate(parts):
                                if re.match(r'^[\d.]+$', part):
                                    s_co = part
                                    break
                
                remark = "; ".join(remark_bits) if remark_bits else ""
    
    except Exception as e:
        return nox, thc, co, pm, pn, s_nox, s_thc, s_co, f"Error: {e}"
    
    return nox, thc, co, pm, pn, s_nox, s_thc, s_co, remark

# ----------- HELPERS --------
def parse_filename(fname: str):
    stem = Path(fname).stem
    parts = stem.split("_")
    test_id = "_".join(parts[:3]) if len(parts) >= 3 else stem
    test_name = "_".join(parts[3:]) if len(parts) > 3 else ""
    try:
        date_val = datetime.fromtimestamp(Path(fname).stat().st_mtime).strftime("%d/%m/%Y")
    except Exception:
        date_val = ""
    return test_id, test_name, date_val

def natural_key(s: str):
    s = str(s)
    parts = re.split(r'(\d+)', s)
    return [int(p) if p.isdigit() else p.lower() for p in parts]

def normalize_header(text: str):
    if text is None:
        return ""
    return re.sub(r'[\s_\-]+', '', str(text).strip().lower())

def parse_excel_for_additional(path: str):
    results = {h: "" for h in ADDITIONAL_HEADERS}
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        normalized_map = {normalize_header(h): h for h in ADDITIONAL_HEADERS}
        remaining = set(normalized_map.keys())
        max_row, max_col = ws.max_row or 0, ws.max_column or 0
        for r in range(1, max_row + 1):
            if not remaining:
                break
            for c in range(1, max_col + 1):
                if not remaining:
                    break
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    continue
                norm = normalize_header(cell.value)
                if norm in remaining:
                    header_name = normalized_map[norm]
                    target_row = r + 2
                    try:
                        val = ws.cell(row=target_row, column=c).value
                        results[header_name] = "" if val is None else str(val).strip()
                    except Exception:
                        results[header_name] = ""
                    remaining.remove(norm)
        return results
    except (InvalidFileException, Exception):
        return results

def parse_txt_for_additional(path: str):
    results = {h: "" for h in ADDITIONAL_HEADERS}
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            lines = [l.strip() for l in f.readlines() if l.strip()]
        normalized_map = {normalize_header(h): h for h in ADDITIONAL_HEADERS}
        remaining = set(normalized_map.keys())
        for i, line in enumerate(lines):
            if not remaining:
                break
            parts = re.split(r'\t|,', line)
            for j, cell in enumerate(parts):
                norm = normalize_header(cell)
                if norm in remaining:
                    header_name = normalized_map[norm]
                    target_index = i + 2
                    if target_index < len(lines):
                        val_parts = re.split(r'\t|,', lines[target_index])
                        if j < len(val_parts):
                            results[header_name] = val_parts[j].strip()
                    remaining.remove(norm)
        return results
    except Exception:
        return results

# ----------- EXCEL WRITER --------
class ExcelWriter:
    def __init__(self, path: str):
        self.path = Path(path)
        if not self.path.exists():
            wb = Workbook()
            ws = wb.active
            ws.append(HEADERS)
            wb.save(self.path)

    def write_all_sorted(self, rows):
        rows.sort(key=lambda r: natural_key(r[1]))
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        for i, row in enumerate(rows, start=1):
            ws.append([i] + row)
        wb.save(self.path)

# ----------- WATCHDOG --------
class FolderHandler(FileSystemEventHandler):
    def __init__(self, app): 
        self.app = app
    def on_created(self, e):
        if not e.is_directory:
            self.app.queue_file(e.src_path)
    def on_modified(self, e):
        if not e.is_directory:
            self.app.queue_file(e.src_path)

# ----------- MODERN UI APP --------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1400x850")
        self.minsize(1000, 600)
        self.resizable(True, True)
        
        # Set dark theme
        self.configure(bg=THEME["bg_dark"])
        
        self.folder = tk.StringVar()
        self.excel = tk.StringVar()
        self.files_queue = []
        self.observer = None
        self.loop_thread = None
        self.stop_flag = threading.Event()
        self.is_monitoring = False
        
        self._build_ui()

    def _build_ui(self):
        """Build modern responsive UI"""
        # Main frame with padding
        main_frame = ttk.Frame(self)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Configure styles
        self._setup_styles()
        
        # ===== TITLE =====
        title_label = ttk.Label(main_frame, text="📊 Summary Sheet Tool v13.0", style='Title.TLabel')
        title_label.pack(anchor="w", pady=(0, 25))
        
        # ===== INPUT SECTION =====
        input_frame = ttk.LabelFrame(main_frame, text=" Configuration ", padding=20)
        input_frame.pack(fill="x", pady=(0, 20))
        input_frame.columnconfigure(1, weight=1)
        
        # Folder
        folder_lbl = ttk.Label(input_frame, text="📁 Reports Folder:", style='Header.TLabel')
        folder_lbl.grid(row=0, column=0, sticky="w", pady=8)
        folder_entry = ttk.Entry(input_frame, textvariable=self.folder, width=60)
        folder_entry.grid(row=0, column=1, sticky="ew", padx=(10, 10))
        ttk.Button(input_frame, text="Browse", command=self.pick_folder, width=12).grid(row=0, column=2, sticky="w")
        
        # Excel
        excel_lbl = ttk.Label(input_frame, text="📄 Excel File:", style='Header.TLabel')
        excel_lbl.grid(row=1, column=0, sticky="w", pady=8)
        excel_entry = ttk.Entry(input_frame, textvariable=self.excel, width=60)
        excel_entry.grid(row=1, column=1, sticky="ew", padx=(10, 10))
        ttk.Button(input_frame, text="Browse", command=self.pick_excel, width=12).grid(row=1, column=2, sticky="w")
        
        # ===== BUTTON SECTION =====
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(0, 20))
        
        self.btn_start = ttk.Button(button_frame, text="▶ Start Monitor", command=self.start_monitor, width=20)
        self.btn_start.pack(side="left", padx=5)
        
        ttk.Button(button_frame, text="⏹ Stop", command=self.stop_monitor, width=20).pack(side="left", padx=5)
        ttk.Button(button_frame, text="🔄 Rescan Now", command=self.rescan, width=20).pack(side="left", padx=5)
        ttk.Button(button_frame, text="📥 Manual Extract", command=self.manual_extract, width=20).pack(side="left", padx=5)
        
        # ===== STATUS SECTION =====
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill="x", pady=(0, 15))
        
        self.status_label = ttk.Label(status_frame, text="⚫ Ready", foreground=THEME["fg_success"])
        self.status_label.pack(side="left", padx=(0, 10))
        
        # ===== LOG SECTION =====
        log_frame = ttk.LabelFrame(main_frame, text=" Activity Log ", padding=10)
        log_frame.pack(fill="both", expand=True)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # Text widget with scrollbar
        txt_frame = ttk.Frame(log_frame)
        txt_frame.grid(row=0, column=0, sticky="nsew")
        txt_frame.columnconfigure(0, weight=1)
        txt_frame.rowconfigure(0, weight=1)
        
        self.txt = tk.Text(
            txt_frame, 
            height=20, 
            width=150,
            bg=THEME["bg_input"],
            fg=THEME["fg_text"],
            insertbackground=THEME["fg_accent"],
            font=('Courier New', 10),
            relief="flat",
            borderwidth=1
        )
        self.txt.grid(row=0, column=0, sticky="nsew")
        
        scrollbar = ttk.Scrollbar(txt_frame, orient="vertical", command=self.txt.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.txt.config(yscrollcommand=scrollbar.set)

    def _setup_styles(self):
        """Setup ttk styles"""
        style = ttk.Style()
        
        # Configure colors
        style.configure('TFrame', background=THEME["bg_dark"])
        style.configure('TLabel', background=THEME["bg_dark"], foreground=THEME["fg_text"])
        style.configure('TEntry', fieldbackground=THEME["bg_input"], foreground=THEME["fg_text"])
        style.configure('TButton', background=THEME["btn_bg"], foreground="white")
        style.configure('TLabelframe', background=THEME["bg_dark"], foreground=THEME["fg_text"])
        style.configure('TLabelframe.Label', background=THEME["bg_dark"], foreground=THEME["fg_accent"])
        
        style.map('TButton', background=[('active', THEME["btn_hover"])])
        style.configure('Header.TLabel', font=('Segoe UI', 11, 'bold'), foreground=THEME["fg_accent"])
        style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'), foreground=THEME["fg_accent"])

    def log(self, msg):
        self.txt.insert("end", f"{datetime.now().strftime('%H:%M:%S')} - {msg}\n")
        self.txt.see("end")
        self.update_idletasks()

    def queue_file(self, path):
        base = os.path.basename(path)
        if base.startswith("~$") or base.startswith("~") or base.endswith(".tmp"):
            return
        if path not in self.files_queue:
            self.files_queue.append(path)

    def _process_file_to_row(self, f):
        test_id, test_name, date = parse_filename(f)
        nox = thc = co = pm = pn = ""
        s_nox = s_thc = s_co = ""
        remark = os.path.basename(f)
        low = f.lower()
        
        if low.endswith(".pdf"):
            nox, thc, co, pm, pn, s_nox, s_thc, s_co, _remark = parse_pdf_comprehensive(f)
        
        add_values = {h: "" for h in ADDITIONAL_HEADERS}
        if low.endswith((".xlsx", ".xlsm", ".xls")):
            add_values = parse_excel_for_additional(f)
        elif low.endswith(".txt"):
            add_values = parse_txt_for_additional(f)
        
        add_list = [add_values.get(h, "") for h in ADDITIONAL_HEADERS]
        return [date, test_id, test_name, nox, thc, co, pm, pn, s_nox, s_thc, s_co, remark] + add_list

    def manual_extract(self):
        if not self.folder.get() or not self.excel.get():
            messagebox.showerror("Error", "Select folder and Excel first.")
            return
        
        self.log("🔍 Manual Extract: scanning folder...")
        self.status_label.config(text="🟡 Processing...")
        self.update_idletasks()
        
        rows = []
        count = 0
        errors = 0
        
        for fname in os.listdir(self.folder.get()):
            path = os.path.join(self.folder.get(), fname)
            if os.path.isfile(path):
                try:
                    row = self._process_file_to_row(path)
                    rows.append(row)
                    count += 1
                    self.log(f"✓ Processed: {os.path.basename(path)}")
                except Exception as e:
                    errors += 1
                    self.log(f"✗ Error: {os.path.basename(path)} -> {e}")
        
        ExcelWriter(self.excel.get()).write_all_sorted(rows)
        self.log(f"✓ Complete. Processed: {count} | Errors: {errors}")
        self.status_label.config(text="🟢 Ready")
        messagebox.showinfo("Success", f"Processed: {count} files\nErrors: {errors}")

    def refresh_sheet_sorted_from_folder(self):
        rows = []
        for fname in os.listdir(self.folder.get()):
            path = os.path.join(self.folder.get(), fname)
            if os.path.isfile(path):
                try: 
                    rows.append(self._process_file_to_row(path))
                except Exception as e: 
                    self.log(f"Error: {fname} -> {e}")
        ExcelWriter(self.excel.get()).write_all_sorted(rows)

    def rescan(self):
        if not self.folder.get() or not self.excel.get():
            messagebox.showerror("Error", "Select folder and Excel first.")
            return
        
        self.log("🔄 Rescanning folder and rebuilding Excel...")
        self.status_label.config(text="🟡 Scanning...")
        self.update_idletasks()
        
        self.refresh_sheet_sorted_from_folder()
        
        self.log("✓ Rescan complete.")
        self.status_label.config(text="🟢 Ready")

    def start_monitor(self):
        if not self.folder.get() or not self.excel.get():
            messagebox.showerror("Error", "Select folder and Excel first.")
            return
        
        self.log("⏳ Initial extraction...")
        self.status_label.config(text="🟡 Extracting...")
        self.update_idletasks()
        
        self.refresh_sheet_sorted_from_folder()
        
        self.log("✓ Extraction complete. Starting monitor...")
        self.observer = Observer()
        self.observer.schedule(FolderHandler(self), self.folder.get(), recursive=False)
        self.observer.start()
        self.stop_flag.clear()
        self.is_monitoring = True
        self.btn_start.config(state="disabled")
        self.status_label.config(text="🟢 Monitoring Active")

        def loop():
            while not self.stop_flag.is_set():
                if self.files_queue:
                    for _ in range(len(self.files_queue)):
                        p = self.files_queue.pop(0)
                        if os.path.isfile(p):
                            try:
                                _ = self._process_file_to_row(p)
                                self.log(f"✓ Processed (queued): {os.path.basename(p)}")
                            except Exception as e:
                                self.log(f"✗ Error processing {os.path.basename(p)} -> {e}")
                    try:
                        self.refresh_sheet_sorted_from_folder()
                        self.log("📊 Excel updated (sorted).")
                    except Exception as e:
                        self.log(f"✗ Error updating Excel: {e}")
                time.sleep(1.0)

        if (self.loop_thread is None) or (not self.loop_thread.is_alive()):
            self.loop_thread = threading.Thread(target=loop, daemon=True)
            self.loop_thread.start()

    def stop_monitor(self):
        if self.observer:
            self.observer.stop()
            self.observer = None
        self.stop_flag.set()
        self.is_monitoring = False
        self.btn_start.config(state="normal")
        self.log("⏹ Monitoring stopped.")
        self.status_label.config(text="🟢 Ready")

    def pick_folder(self):
        d = filedialog.askdirectory()
        if d: 
            self.folder.set(d)
            self.log(f"✓ Folder: {d}")

    def pick_excel(self):
        f = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if f: 
            self.excel.set(f)
            self.log(f"✓ Excel: {f}")

if __name__ == "__main__":
    app = App()
    app.mainloop()
