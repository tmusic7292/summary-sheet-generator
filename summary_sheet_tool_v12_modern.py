import os
os.environ["NUMBA_DISABLE_JIT"] = "1"

#!/usr/bin/env python3
"""
Summary Sheet Tool v12.0 — Modern PyQt5 UI
Features:
- Real-time extraction via watchdog
- FIXED: Separate NOx & THC extraction from Specific Results section
- Modern Material Design UI (PyQt5)
- Stunning dark theme with neon accents
- All previous features retained
"""

import os
import re
import time
import threading
from pathlib import Path
from datetime import datetime
import pdfplumber
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# PyQt5 imports
import sys
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QLabel, QLineEdit, QPushButton, QTextEdit, QFileDialog,
                             QMessageBox, QFrame, QProgressBar)
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QObject, QThread
from PyQt5.QtGui import QFont, QIcon, QColor, QPixmap
from PyQt5.QtCore import QSize

# --- CONFIG ---
APP_TITLE = "Summary Sheet Tool v12.0 — Modern Edition"
ADDITIONAL_HEADERS = [
    "Injector 1", "Injector 2", "Injector 3", "Injector 4",
    "DOC", "DPF", "Engine Number", "Dataset"
]
# Updated headers with separate NOx and THC columns
HEADERS = ["S No.", "Date", "Test ID", "Test Name", "NOx", "THC", "CO", "PM", "PN", "Remark"] + ADDITIONAL_HEADERS

# --- COLOR THEME (Modern Material Design) ---
THEME = {
    "bg_dark": "#0f0f0f",
    "bg_card": "#1a1a1a",
    "bg_input": "#2a2a2a",
    "fg_text": "#e0e0e0",
    "fg_secondary": "#b0b0b0",
    "fg_accent": "#00d4ff",
    "fg_success": "#00ff88",
    "fg_error": "#ff4466",
    "btn_primary": "#0066ff",
    "btn_hover": "#0080ff",
    "btn_accent": "#00d4ff",
    "border": "#404040",
}

# ========== PDF PARSING - FIXED NOx/THC EXTRACTION ==========
def parse_pdf_fixed(path: str):
    """
    Extract ONLY from Specific Results section:
    - NOx [g/kWh]
    - THC [g/kWh]
    - CO (COL - Tailpipe) [g/kWh]
    - PM, PN from Emission Limits if available
    """
    nox = ""
    thc = ""
    co = ""
    pm = ""
    pn = ""
    remark = ""
    remark_bits = []
    
    try:
        with pdfplumber.open(path) as pdf:
            for page_idx, page in enumerate(pdf.pages):
                txt = page.extract_text() or ""
                lines = txt.splitlines()
                
                # ===== SECTION 1: Specific Results (TAILPIPE) =====
                # Look for "Specific results" section
                in_specific_results = False
                for i, line in enumerate(lines):
                    # Check if we're entering Specific Results section
                    if "Specific results" in line and "Tailpipe" in line:
                        in_specific_results = True
                        continue
                    
                    # Exit specific results when we hit another section
                    if in_specific_results and any(x in line for x in ["Mass results", "Fuel consumption", "Engine performance"]):
                        in_specific_results = False
                    
                    if in_specific_results:
                        # Extract NOx value
                        if "NOX [g/kWh]" in line or "NOx [g/kWh]" in line:
                            # Look for number in this line or next lines
                            nums = re.findall(r"([0-9]+\.?[0-9]*)", line)
                            if nums and line.split()[-1].replace('.', '').replace('-', '').isdigit():
                                nox = line.split()[-1]
                        
                        # Extract THC value
                        if "THC [g/kWh]" in line:
                            nums = re.findall(r"([0-9]+\.?[0-9]*)", line)
                            if nums and line.split()[-1].replace('.', '').replace('-', '').isdigit():
                                thc = line.split()[-1]
                        
                        # Extract CO (COL - Tailpipe) value
                        if "CO (COL - Tailpipe)" in line or "CO (COL-Tailpipe)" in line:
                            nums = re.findall(r"([0-9]+\.?[0-9]*)", line)
                            if nums and line.split()[-1].replace('.', '').replace('-', '').isdigit():
                                co = line.split()[-1]
                
                # ===== SECTION 2: Emission Limits (for PM, PN) =====
                if "Emission limits - Stage V NRE" in txt:
                    for l in lines:
                        nums = re.findall(r"([0-9][0-9Ee\+\-\.]*)", l)
                        
                        if "Particulate mass" in l and nums:
                            pm = nums[0] if nums else ""
                        elif "Particulate number" in l and nums:
                            pn = nums[0] if nums else ""
                        
                        if ("Passed" in l) or ("Failed" in l):
                            remark_bits.append(l.strip())
                
                remark = "; ".join(remark_bits) if remark_bits else ""
    
    except Exception as e:
        return "", "", "", "", "", f"PDF parse error: {e}"
    
    return nox, thc, co, pm, pn, remark

# ----------- HELPERS --------
def parse_filename(fname: str):
    stem = Path(fname).stem
    parts = stem.split("_")
    test_id = "_".join(parts[:3]) if len(parts) >= 3 else stem
    test_name = "_".join(parts[3:]) if len(parts) > 3 else ""
    try:
        date_val = datetime.fromtimestamp(Path(fname).stat().st_mtime).strftime("%d/%m/%Y")
    except Exception:
        date_val = ""
    return test_id, test_name, date_val

def natural_key(s: str):
    s = str(s)
    parts = re.split(r'(\d+)', s)
    return [int(p) if p.isdigit() else p.lower() for p in parts]

def normalize_header(text: str):
    if text is None:
        return ""
    return re.sub(r'[\s_\-]+', '', str(text).strip().lower())

def parse_excel_for_additional(path: str):
    results = {h: "" for h in ADDITIONAL_HEADERS}
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        normalized_map = {normalize_header(h): h for h in ADDITIONAL_HEADERS}
        remaining = set(normalized_map.keys())
        max_row, max_col = ws.max_row or 0, ws.max_column or 0
        for r in range(1, max_row + 1):
            if not remaining:
                break
            for c in range(1, max_col + 1):
                if not remaining:
                    break
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    continue
                norm = normalize_header(cell.value)
                if norm in remaining:
                    header_name = normalized_map[norm]
                    target_row = r + 2
                    try:
                        val = ws.cell(row=target_row, column=c).value
                        results[header_name] = "" if val is None else str(val).strip()
                    except Exception:
                        results[header_name] = ""
                    remaining.remove(norm)
        return results
    except (InvalidFileException, Exception):
        return results

def parse_txt_for_additional(path: str):
    results = {h: "" for h in ADDITIONAL_HEADERS}
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            lines = [l.strip() for l in f.readlines() if l.strip()]
        normalized_map = {normalize_header(h): h for h in ADDITIONAL_HEADERS}
        remaining = set(normalized_map.keys())
        for i, line in enumerate(lines):
            if not remaining:
                break
            parts = re.split(r'\t|,', line)
            for j, cell in enumerate(parts):
                norm = normalize_header(cell)
                if norm in remaining:
                    header_name = normalized_map[norm]
                    target_index = i + 2
                    if target_index < len(lines):
                        val_parts = re.split(r'\t|,', lines[target_index])
                        if j < len(val_parts):
                            results[header_name] = val_parts[j].strip()
                    remaining.remove(norm)
        return results
    except Exception:
        return results

# ----------- EXCEL WRITER --------
class ExcelWriter:
    def __init__(self, path: str):
        self.path = Path(path)
        if not self.path.exists():
            wb = Workbook()
            ws = wb.active
            ws.append(HEADERS)
            wb.save(self.path)

    def write_all_sorted(self, rows):
        rows.sort(key=lambda r: natural_key(r[1]))
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        for i, row in enumerate(rows, start=1):
            ws.append([i] + row)
        wb.save(self.path)

# ----------- WATCHDOG --------
class FolderHandler(FileSystemEventHandler):
    def __init__(self, callback):
        self.callback = callback
    
    def on_created(self, e):
        if not e.is_directory:
            self.callback(e.src_path)
    
    def on_modified(self, e):
        if not e.is_directory:
            self.callback(e.src_path)

# ----------- WORKER THREAD --------
class WorkerThread(QThread):
    log_signal = pyqtSignal(str)
    status_signal = pyqtSignal(str)
    
    def __init__(self, folder, excel):
        super().__init__()
        self.folder = folder
        self.excel = excel
        self.running = True
        self.files_queue = []
        self.observer = None
    
    def queue_file(self, path):
        base = os.path.basename(path)
        if base.startswith("~$") or base.startswith("~") or base.endswith(".tmp"):
            return
        if path not in self.files_queue:
            self.files_queue.append(path)
    
    def _process_file(self, f):
        test_id, test_name, date = parse_filename(f)
        low = f.lower()
        
        if low.endswith(".pdf"):
            nox, thc, co, pm, pn, _remark = parse_pdf_fixed(f)
        else:
            nox = thc = co = pm = pn = ""
        
        remark = os.path.basename(f)
        add_values = {h: "" for h in ADDITIONAL_HEADERS}
        
        if low.endswith((".xlsx", ".xlsm", ".xls")):
            add_values = parse_excel_for_additional(f)
        elif low.endswith(".txt"):
            add_values = parse_txt_for_additional(f)
        
        add_list = [add_values.get(h, "") for h in ADDITIONAL_HEADERS]
        return [date, test_id, test_name, nox, thc, co, pm, pn, remark] + add_list
    
    def refresh_sheet(self):
        rows = []
        for fname in os.listdir(self.folder):
            path = os.path.join(self.folder, fname)
            if os.path.isfile(path):
                try:
                    rows.append(self._process_file(path))
                except Exception as e:
                    self.log_signal.emit(f"❌ Error: {fname} -> {e}")
        
        ExcelWriter(self.excel).write_all_sorted(rows)
    
    def run(self):
        self.log_signal.emit("⏳ Initial extraction...")
        self.status_signal.emit("🟡 Extracting...")
        
        try:
            self.refresh_sheet()
            self.log_signal.emit("✓ Extraction complete. Starting monitor...")
            self.status_signal.emit("🟢 Monitoring Active")
            
            self.observer = Observer()
            self.observer.schedule(FolderHandler(self.queue_file), self.folder, recursive=False)
            self.observer.start()
            
            while self.running:
                if self.files_queue:
                    for _ in range(len(self.files_queue)):
                        p = self.files_queue.pop(0)
                        if os.path.isfile(p):
                            try:
                                _ = self._process_file(p)
                                self.log_signal.emit(f"✓ Processed: {os.path.basename(p)}")
                            except Exception as e:
                                self.log_signal.emit(f"❌ Error: {os.path.basename(p)} -> {e}")
                    
                    try:
                        self.refresh_sheet()
                        self.log_signal.emit("📊 Excel updated")
                    except Exception as e:
                        self.log_signal.emit(f"❌ Error updating Excel: {e}")
                
                time.sleep(1.0)
        
        except Exception as e:
            self.log_signal.emit(f"❌ Monitor error: {e}")
            self.status_signal.emit("🔴 Error")
    
    def stop(self):
        self.running = False
        if self.observer:
            self.observer.stop()

# ========== MODERN PYQT5 UI ==========
class ModernApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.setGeometry(100, 100, 1200, 800)
        self.setStyleSheet(self._get_stylesheet())
        
        self.folder = ""
        self.excel = ""
        self.worker_thread = None
        
        self._build_ui()
        self.setWindowIcon(self._create_icon())
    
    def _get_stylesheet(self):
        return f"""
        QMainWindow {{
            background-color: {THEME['bg_dark']};
            color: {THEME['fg_text']};
        }}
        
        QWidget {{
            background-color: {THEME['bg_dark']};
            color: {THEME['fg_text']};
        }}
        
        QLineEdit {{
            background-color: {THEME['bg_input']};
            border: 2px solid {THEME['border']};
            border-radius: 8px;
            padding: 10px;
            color: {THEME['fg_text']};
            selection-background-color: {THEME['fg_accent']};
            font-size: 11px;
        }}
        
        QLineEdit:focus {{
            border: 2px solid {THEME['fg_accent']};
        }}
        
        QPushButton {{
            background-color: {THEME['btn_primary']};
            color: white;
            border: none;
            border-radius: 8px;
            padding: 12px 24px;
            font-weight: bold;
            font-size: 11px;
        }}
        
        QPushButton:hover {{
            background-color: {THEME['btn_hover']};
        }}
        
        QPushButton:pressed {{
            background-color: {THEME['btn_accent']};
        }}
        
        QPushButton:disabled {{
            background-color: #404040;
            color: #808080;
        }}
        
        QTextEdit {{
            background-color: {THEME['bg_input']};
            border: 2px solid {THEME['border']};
            border-radius: 8px;
            padding: 10px;
            color: {THEME['fg_text']};
            font-family: 'Courier New';
            font-size: 10px;
        }}
        
        QLabel {{
            color: {THEME['fg_text']};
        }}
        
        QFrame {{
            background-color: {THEME['bg_card']};
            border-radius: 12px;
        }}
        
        QProgressBar {{
            background-color: {THEME['bg_input']};
            border: 2px solid {THEME['border']};
            border-radius: 8px;
            height: 8px;
        }}
        
        QProgressBar::chunk {{
            background-color: {THEME['fg_accent']};
            border-radius: 6px;
        }}
        """
    
    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout()
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(30, 30, 30, 30)
        
        # ===== TITLE =====
        title = QLabel("📊 Summary Sheet Tool v12.0")
        title_font = QFont("Segoe UI", 18, QFont.Bold)
        title.setFont(title_font)
        title.setStyleSheet(f"color: {THEME['fg_accent']};")
        main_layout.addWidget(title)
        
        # ===== CONFIG SECTION =====
        config_frame = QFrame()
        config_frame.setStyleSheet(f"background-color: {THEME['bg_card']}; border-radius: 12px; padding: 20px;")
        config_layout = QVBoxLayout()
        config_layout.setSpacing(15)
        
        # Folder
        folder_layout = QHBoxLayout()
        folder_label = QLabel("📁 Reports Folder:")
        folder_label.setStyleSheet(f"color: {THEME['fg_accent']}; font-weight: bold;")
        folder_label.setMinimumWidth(150)
        self.folder_input = QLineEdit()
        self.folder_input.setPlaceholderText("Select your reports folder...")
        folder_btn = QPushButton("Browse")
        folder_btn.setMaximumWidth(100)
        folder_btn.clicked.connect(self.pick_folder)
        folder_layout.addWidget(folder_label)
        folder_layout.addWidget(self.folder_input)
        folder_layout.addWidget(folder_btn)
        config_layout.addLayout(folder_layout)
        
        # Excel
        excel_layout = QHBoxLayout()
        excel_label = QLabel("📄 Excel File:")
        excel_label.setStyleSheet(f"color: {THEME['fg_accent']}; font-weight: bold;")
        excel_label.setMinimumWidth(150)
        self.excel_input = QLineEdit()
        self.excel_input.setPlaceholderText("Select or create Excel file...")
        excel_btn = QPushButton("Browse")
        excel_btn.setMaximumWidth(100)
        excel_btn.clicked.connect(self.pick_excel)
        excel_layout.addWidget(excel_label)
        excel_layout.addWidget(self.excel_input)
        excel_layout.addWidget(excel_btn)
        config_layout.addLayout(excel_layout)
        
        config_frame.setLayout(config_layout)
        main_layout.addWidget(config_frame)
        
        # ===== BUTTONS SECTION =====
        button_frame = QFrame()
        button_frame.setStyleSheet(f"background-color: {THEME['bg_card']}; border-radius: 12px;")
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        button_layout.setContentsMargins(20, 15, 20, 15)
        
        self.btn_start = QPushButton("▶ Start Monitor")
        self.btn_start.clicked.connect(self.start_monitor)
        self.btn_start.setMinimumHeight(45)
        
        self.btn_stop = QPushButton("⏹ Stop")
        self.btn_stop.clicked.connect(self.stop_monitor)
        self.btn_stop.setMinimumHeight(45)
        self.btn_stop.setEnabled(False)
        
        btn_rescan = QPushButton("🔄 Rescan")
        btn_rescan.clicked.connect(self.rescan)
        btn_rescan.setMinimumHeight(45)
        
        btn_manual = QPushButton("📥 Manual Extract")
        btn_manual.clicked.connect(self.manual_extract)
        btn_manual.setMinimumHeight(45)
        
        button_layout.addWidget(self.btn_start)
        button_layout.addWidget(self.btn_stop)
        button_layout.addWidget(btn_rescan)
        button_layout.addWidget(btn_manual)
        
        button_frame.setLayout(button_layout)
        main_layout.addWidget(button_frame)
        
        # ===== STATUS BAR =====
        status_frame = QFrame()
        status_frame.setStyleSheet(f"background-color: {THEME['bg_card']}; border-radius: 12px; padding: 15px;")
        status_layout = QHBoxLayout()
        
        self.status_label = QLabel("⚫ Ready")
        self.status_label.setStyleSheet(f"color: {THEME['fg_success']}; font-weight: bold; font-size: 12px;")
        status_layout.addWidget(self.status_label)
        status_layout.addStretch()
        
        status_frame.setLayout(status_layout)
        main_layout.addWidget(status_frame)
        
        # ===== LOG SECTION =====
        log_label = QLabel("📋 Activity Log")
        log_label.setStyleSheet(f"color: {THEME['fg_accent']}; font-weight: bold; font-size: 12px;")
        main_layout.addWidget(log_label)
        
        log_frame = QFrame()
        log_frame.setStyleSheet(f"background-color: {THEME['bg_card']}; border-radius: 12px; padding: 10px;")
        log_layout = QVBoxLayout()
        log_layout.setContentsMargins(0, 0, 0, 0)
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMinimumHeight(300)
        log_layout.addWidget(self.log_text)
        log_frame.setLayout(log_layout)
        main_layout.addWidget(log_frame)
        
        central.setLayout(main_layout)
    
    def _create_icon(self):
        """Create a simple icon"""
        pixmap = QPixmap(64, 64)
        pixmap.fill(QColor(THEME['bg_dark']))
        return QIcon(pixmap)
    
    def log(self, msg):
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.log_text.append(f"[{timestamp}] {msg}")
    
    def pick_folder(self):
        d = QFileDialog.getExistingDirectory(self, "Select Reports Folder")
        if d:
            self.folder = d
            self.folder_input.setText(d)
            self.log(f"✓ Folder: {d}")
    
    def pick_excel(self):
        f, _ = QFileDialog.getSaveFileName(self, "Select Excel File", "", "Excel (*.xlsx)")
        if f:
            self.excel = f
            self.excel_input.setText(f)
            self.log(f"✓ Excel: {f}")
    
    def start_monitor(self):
        if not self.folder or not self.excel:
            QMessageBox.critical(self, "Error", "Select folder and Excel file first!")
            return
        
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)
        
        self.worker_thread = WorkerThread(self.folder, self.excel)
        self.worker_thread.log_signal.connect(self.log)
        self.worker_thread.status_signal.connect(self.update_status)
        self.worker_thread.start()
    
    def stop_monitor(self):
        if self.worker_thread:
            self.worker_thread.stop()
            self.worker_thread.wait()
        
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.update_status("🟢 Ready")
        self.log("⏹ Monitoring stopped")
    
    def rescan(self):
        if not self.folder or not self.excel:
            QMessageBox.critical(self, "Error", "Select folder and Excel file first!")
            return
        
        self.log("🔄 Rescanning...")
        self.update_status("🟡 Scanning...")
        
        try:
            rows = []
            for fname in os.listdir(self.folder):
                path = os.path.join(self.folder, fname)
                if os.path.isfile(path):
                    try:
                        test_id, test_name, date = parse_filename(path)
                        low = path.lower()
                        
                        if low.endswith(".pdf"):
                            nox, thc, co, pm, pn, _remark = parse_pdf_fixed(path)
                        else:
                            nox = thc = co = pm = pn = ""
                        
                        remark = os.path.basename(path)
                        add_values = {h: "" for h in ADDITIONAL_HEADERS}
                        
                        if low.endswith((".xlsx", ".xlsm", ".xls")):
                            add_values = parse_excel_for_additional(path)
                        elif low.endswith(".txt"):
                            add_values = parse_txt_for_additional(path)
                        
                        add_list = [add_values.get(h, "") for h in ADDITIONAL_HEADERS]
                        rows.append([date, test_id, test_name, nox, thc, co, pm, pn, remark] + add_list)
                    except Exception as e:
                        self.log(f"❌ Error: {fname} -> {e}")
            
            ExcelWriter(self.excel).write_all_sorted(rows)
            self.log("✓ Rescan complete")
            self.update_status("🟢 Ready")
        except Exception as e:
            self.log(f"❌ Rescan error: {e}")
            self.update_status("🔴 Error")
    
    def manual_extract(self):
        if not self.folder or not self.excel:
            QMessageBox.critical(self, "Error", "Select folder and Excel file first!")
            return
        
        self.log("📥 Manual extract starting...")
        self.update_status("🟡 Processing...")
        
        try:
            rows = []
            count = 0
            for fname in os.listdir(self.folder):
                path = os.path.join(self.folder, fname)
                if os.path.isfile(path):
                    try:
                        test_id, test_name, date = parse_filename(path)
                        low = path.lower()
                        
                        if low.endswith(".pdf"):
                            nox, thc, co, pm, pn, _remark = parse_pdf_fixed(path)
                        else:
                            nox = thc = co = pm = pn = ""
                        
                        remark = os.path.basename(path)
                        add_values = {h: "" for h in ADDITIONAL_HEADERS}
                        
                        if low.endswith((".xlsx", ".xlsm", ".xls")):
                            add_values = parse_excel_for_additional(path)
                        elif low.endswith(".txt"):
                            add_values = parse_txt_for_additional(path)
                        
                        add_list = [add_values.get(h, "") for h in ADDITIONAL_HEADERS]
                        rows.append([date, test_id, test_name, nox, thc, co, pm, pn, remark] + add_list)
                        count += 1
                        self.log(f"✓ {os.path.basename(path)}")
                    except Exception as e:
                        self.log(f"❌ {os.path.basename(path)}: {e}")
            
            ExcelWriter(self.excel).write_all_sorted(rows)
            self.log(f"✓ Manual extract complete - {count} files")
            self.update_status("🟢 Ready")
            QMessageBox.information(self, "Success", f"Extracted {count} files successfully!")
        except Exception as e:
            self.log(f"❌ Error: {e}")
            self.update_status("🔴 Error")
    
    def update_status(self, status):
        self.status_label.setText(status)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ModernApp()
    window.show()
    sys.exit(app.exec_())
