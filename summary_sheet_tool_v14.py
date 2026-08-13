import os
os.environ["NUMBA_DISABLE_JIT"] = "1"

#!/usr/bin/env python3
"""
Summary Sheet Tool v14.0 — Enhanced UI + Fixed PDF Extraction
Features:
- Real-time extraction via watchdog
- Correct PDF extraction from Emission Limits & Specific Results
- NOx+HC, THC, CO, PM, PN, S_Nox, S_THC extraction
- Enhanced UI with better text visibility and layout
- Responsive modern design
"""

import os
import re
import time
import threading
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pdfplumber
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# --- CONFIG ---
APP_TITLE = "Summary Sheet Tool v14.0 — Enhanced Edition"
ADDITIONAL_HEADERS = [
    "Injector 1", "Injector 2", "Injector 3", "Injector 4",
    "DOC", "DPF", "Engine Number", "Dataset"
]
# Updated headers with correct column order
HEADERS = ["S No.", "Date", "Test ID", "Test Name", "NOx+HC", "S_Nox", "THC", "S_THC", "CO", "PM", "PN", "Remark"] + ADDITIONAL_HEADERS

# --- ENHANCED COLOR THEME ---
THEME = {
    "bg_dark": "#0d0d0d",
    "bg_card": "#1a1a1a",
    "bg_input": "#252525",
    "fg_text": "#f5f5f5",
    "fg_secondary": "#c0c0c0",
    "fg_accent": "#00d9ff",
    "fg_success": "#00ff99",
    "fg_error": "#ff5577",
    "btn_primary": "#0066ff",
    "btn_hover": "#0080ff",
    "btn_text": "#ffffff",
    "border": "#333333",
}

# ========== ENHANCED PDF PARSING ==========
def parse_pdf_comprehensive(path: str):
    """
    Extract data from:
    1. Emission Limits → NOx+HC, CO, PM, PN
    2. Specific Results → S_Nox, THC (S_THC)
    """
    # Emission Limits extraction
    nox_hc = ""
    co = ""
    pm = ""
    pn = ""
    remark = ""
    remark_bits = []
    
    # Specific Results extraction
    s_nox = ""
    s_thc = ""
    
    try:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                txt = page.extract_text() or ""
                lines = txt.splitlines()
                
                # ===== SECTION 1: Emission Limits =====
                if "Emission limits - Stage V NRE" in txt:
                    for l in lines:
                        # Extract NOx+HC
                        if "NOX+HC" in l or "NOx+HC" in l:
                            parts = l.split()
                            for part in parts:
                                if re.match(r'^[\d.]+$', part):
                                    nox_hc = part
                                    break
                        
                        # Extract CO uncorrected
                        if "CO uncorrected" in l:
                            parts = l.split()
                            for part in parts:
                                if re.match(r'^[\d.]+$', part):
                                    co = part
                                    break
                        
                        # Extract Particulate mass
                        if "Particulate mass" in l:
                            parts = l.split()
                            for part in parts:
                                if re.match(r'^[\d.eE+\-]+$', part):
                                    pm = part
                                    break
                        
                        # Extract Particulate number
                        if "Particulate number" in l:
                            parts = l.split()
                            for part in parts:
                                if re.match(r'^[\d.eE+\-]+$', part):
                                    pn = part
                                    break
                        
                        # Get remark from Pass/Fail status
                        if ("Passed" in l) or ("Failed" in l):
                            remark_bits.append(l.strip())
                
                # ===== SECTION 2: Specific Results =====
                if "Specific results" in txt.lower() and "Tailpipe" in txt:
                    in_specific = False
                    for i, line in enumerate(lines):
                        if "Specific results" in line.lower():
                            in_specific = True
                            continue
                        
                        if in_specific:
                            # Stop at next section
                            if any(x in line for x in ["Mass results", "Fuel consumption", "Engine performance"]):
                                in_specific = False
                            
                            # Extract S_Nox from "NOX [g/kWh]"
                            if re.search(r'\bNOX\s*\[g/kWh\]', line, re.IGNORECASE) or re.search(r'\bNOx\s*\[g/kWh\]', line, re.IGNORECASE):
                                parts = line.split()
                                for part in parts:
                                    if re.match(r'^[\d.]+$', part):
                                        s_nox = part
                                        break
                            
                            # Extract S_THC from "THC [g/kWh]"
                            if re.search(r'\bTHC\s*\[g/kWh\]', line, re.IGNORECASE):
                                parts = line.split()
                                for part in parts:
                                    if re.match(r'^[\d.]+$', part):
                                        s_thc = part
                                        break
                
                remark = "; ".join(remark_bits) if remark_bits else ""
    
    except Exception as e:
        return nox_hc, co, pm, pn, s_nox, s_thc, f"Error: {e}"
    
    return nox_hc, co, pm, pn, s_nox, s_thc, remark

# ----------- HELPERS --------
def parse_filename(fname: str):
    stem = Path(fname).stem
    parts = stem.split("_")
    test_id = "_".join(parts[:3]) if len(parts) >= 3 else stem
    test_name = "_".join(parts[3:]) if len(parts) > 3 else ""
    try:
        date_val = datetime.fromtimestamp(Path(fname).stat().st_mtime).strftime("%d/%m/%Y")
    except Exception:
        date_val = ""
    return test_id, test_name, date_val

def natural_key(s: str):
    s = str(s)
    parts = re.split(r'(\d+)', s)
    return [int(p) if p.isdigit() else p.lower() for p in parts]

def normalize_header(text: str):
    if text is None:
        return ""
    return re.sub(r'[\s_\-]+', '', str(text).strip().lower())

def parse_excel_for_additional(path: str):
    results = {h: "" for h in ADDITIONAL_HEADERS}
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        normalized_map = {normalize_header(h): h for h in ADDITIONAL_HEADERS}
        remaining = set(normalized_map.keys())
        max_row, max_col = ws.max_row or 0, ws.max_column or 0
        for r in range(1, max_row + 1):
            if not remaining:
                break
            for c in range(1, max_col + 1):
                if not remaining:
                    break
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    continue
                norm = normalize_header(cell.value)
                if norm in remaining:
                    header_name = normalized_map[norm]
                    target_row = r + 2
                    try:
                        val = ws.cell(row=target_row, column=c).value
                        results[header_name] = "" if val is None else str(val).strip()
                    except Exception:
                        results[header_name] = ""
                    remaining.remove(norm)
        return results
    except (InvalidFileException, Exception):
        return results

def parse_txt_for_additional(path: str):
    results = {h: "" for h in ADDITIONAL_HEADERS}
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            lines = [l.strip() for l in f.readlines() if l.strip()]
        normalized_map = {normalize_header(h): h for h in ADDITIONAL_HEADERS}
        remaining = set(normalized_map.keys())
        for i, line in enumerate(lines):
            if not remaining:
                break
            parts = re.split(r'\t|,', line)
            for j, cell in enumerate(parts):
                norm = normalize_header(cell)
                if norm in remaining:
                    header_name = normalized_map[norm]
                    target_index = i + 2
                    if target_index < len(lines):
                        val_parts = re.split(r'\t|,', lines[target_index])
                        if j < len(val_parts):
                            results[header_name] = val_parts[j].strip()
                    remaining.remove(norm)
        return results
    except Exception:
        return results

# ----------- EXCEL WRITER --------
class ExcelWriter:
    def __init__(self, path: str):
        self.path = Path(path)
        if not self.path.exists():
            wb = Workbook()
            ws = wb.active
            ws.append(HEADERS)
            wb.save(self.path)

    def write_all_sorted(self, rows):
        rows.sort(key=lambda r: natural_key(r[1]))
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        for i, row in enumerate(rows, start=1):
            ws.append([i] + row)
        wb.save(self.path)

# ----------- WATCHDOG --------
class FolderHandler(FileSystemEventHandler):
    def __init__(self, app): 
        self.app = app
    def on_created(self, e):
        if not e.is_directory:
            self.app.queue_file(e.src_path)
    def on_modified(self, e):
        if not e.is_directory:
            self.app.queue_file(e.src_path)

# ========== ENHANCED MODERN UI ==========
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1500x950")
        self.minsize(1200, 750)
        self.resizable(True, True)
        
        # Set dark theme
        self.configure(bg=THEME["bg_dark"])
        
        self.folder = tk.StringVar()
        self.excel = tk.StringVar()
        self.files_queue = []
        self.observer = None
        self.loop_thread = None
        self.stop_flag = threading.Event()
        self.is_monitoring = False
        
        self._setup_styles()
        self._build_ui()

    def _setup_styles(self):
        """Setup enhanced ttk styles with better visibility"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure colors with better contrast
        style.configure('TFrame', background=THEME["bg_dark"])
        style.configure('TLabel', background=THEME["bg_dark"], foreground=THEME["fg_text"], font=('Segoe UI', 11))
        style.configure('TEntry', fieldbackground=THEME["bg_input"], foreground=THEME["fg_text"], font=('Segoe UI', 11))
        style.configure('TButton', background=THEME["btn_primary"], foreground=THEME["btn_text"], font=('Segoe UI', 11, 'bold'), borderwidth=0, focuscolor='none')
        style.configure('TLabelframe', background=THEME["bg_card"], foreground=THEME["fg_text"], font=('Segoe UI', 12, 'bold'))
        style.configure('TLabelframe.Label', background=THEME["bg_card"], foreground=THEME["fg_accent"], font=('Segoe UI', 12, 'bold'))
        
        style.map('TButton', 
                 background=[('active', THEME["btn_hover"]), ('pressed', THEME["fg_accent"])],
                 foreground=[('active', THEME["btn_text"])])
        
        style.configure('Header.TLabel', font=('Segoe UI', 12, 'bold'), foreground=THEME["fg_accent"])
        style.configure('Title.TLabel', font=('Segoe UI', 20, 'bold'), foreground=THEME["fg_accent"])
        style.configure('Status.TLabel', font=('Segoe UI', 12, 'bold'), foreground=THEME["fg_success"])

    def _build_ui(self):
        """Build enhanced responsive UI with better visibility"""
        # Main frame with generous padding
        main_frame = ttk.Frame(self)
        main_frame.pack(fill="both", expand=True, padx=25, pady=25)
        main_frame.columnconfigure(0, weight=1)
        
        # ===== TITLE =====
        title_label = ttk.Label(main_frame, text="📊 Summary Sheet Tool v14.0", style='Title.TLabel')
        title_label.pack(anchor="w", pady=(0, 30))
        
        # ===== CONFIG SECTION =====
        config_frame = ttk.LabelFrame(main_frame, text=" Configuration ", padding=25)
        config_frame.pack(fill="x", pady=(0, 25))
        config_frame.columnconfigure(1, weight=1)
        
        # Folder row
        folder_lbl = ttk.Label(config_frame, text="📁 Reports Folder:", style='Header.TLabel')
        folder_lbl.grid(row=0, column=0, sticky="w", pady=15, padx=(0, 20))
        folder_entry = ttk.Entry(config_frame, textvariable=self.folder, width=70)
        folder_entry.grid(row=0, column=1, sticky="ew", padx=(0, 15), ipady=8)
        ttk.Button(config_frame, text="Browse", command=self.pick_folder, width=15).grid(row=0, column=2, sticky="ew", padx=5)
        
        # Excel row
        excel_lbl = ttk.Label(config_frame, text="📄 Excel File:", style='Header.TLabel')
        excel_lbl.grid(row=1, column=0, sticky="w", pady=15, padx=(0, 20))
        excel_entry = ttk.Entry(config_frame, textvariable=self.excel, width=70)
        excel_entry.grid(row=1, column=1, sticky="ew", padx=(0, 15), ipady=8)
        ttk.Button(config_frame, text="Browse", command=self.pick_excel, width=15).grid(row=1, column=2, sticky="ew", padx=5)
        
        # ===== BUTTON SECTION =====
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(0, 25))
        button_frame.columnconfigure(0, weight=1)
        
        btn_container = ttk.Frame(button_frame)
        btn_container.pack(fill="x")
        
        self.btn_start = ttk.Button(btn_container, text="▶  Start Monitor", command=self.start_monitor, width=22)
        self.btn_start.pack(side="left", padx=8, ipady=12)
        
        ttk.Button(btn_container, text="⏹  Stop", command=self.stop_monitor, width=22).pack(side="left", padx=8, ipady=12)
        ttk.Button(btn_container, text="🔄  Rescan Now", command=self.rescan, width=22).pack(side="left", padx=8, ipady=12)
        ttk.Button(btn_container, text="📥  Manual Extract", command=self.manual_extract, width=22).pack(side="left", padx=8, ipady=12)
        
        # ===== STATUS SECTION =====
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill="x", pady=(0, 20))
        
        self.status_label = ttk.Label(status_frame, text="⚫ Ready", style='Status.TLabel')
        self.status_label.pack(side="left", padx=10, pady=10)
        
        # ===== LOG SECTION =====
        log_label = ttk.Label(main_frame, text="📋 Activity Log", style='Header.TLabel')
        log_label.pack(anchor="w", pady=(0, 10))
        
        log_frame = ttk.Frame(main_frame)
        log_frame.pack(fill="both", expand=True)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # Text widget with scrollbar
        self.txt = tk.Text(
            log_frame, 
            height=25, 
            width=180,
            bg=THEME["bg_input"],
            fg=THEME["fg_text"],
            insertbackground=THEME["fg_accent"],
            font=('Courier New', 11),
            relief="flat",
            borderwidth=1,
            padx=15,
            pady=12
        )
        self.txt.grid(row=0, column=0, sticky="nsew")
        
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.txt.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.txt.config(yscrollcommand=scrollbar.set)

    def log(self, msg):
        self.txt.insert("end", f"{datetime.now().strftime('%H:%M:%S')} - {msg}\n")
        self.txt.see("end")
        self.update_idletasks()

    def queue_file(self, path):
        base = os.path.basename(path)
        if base.startswith("~$") or base.startswith("~") or base.endswith(".tmp"):
            return
        if path not in self.files_queue:
            self.files_queue.append(path)

    def _process_file_to_row(self, f):
        test_id, test_name, date = parse_filename(f)
        nox_hc = co = pm = pn = s_nox = s_thc = ""
        remark = os.path.basename(f)
        low = f.lower()
        
        if low.endswith(".pdf"):
            nox_hc, co, pm, pn, s_nox, s_thc, _remark = parse_pdf_comprehensive(f)
        
        add_values = {h: "" for h in ADDITIONAL_HEADERS}
        if low.endswith((".xlsx", ".xlsm", ".xls")):
            add_values = parse_excel_for_additional(f)
        elif low.endswith(".txt"):
            add_values = parse_txt_for_additional(f)
        
        add_list = [add_values.get(h, "") for h in ADDITIONAL_HEADERS]
        return [date, test_id, test_name, nox_hc, s_nox, s_thc, s_thc, co, pm, pn, remark] + add_list

    def manual_extract(self):
        if not self.folder.get() or not self.excel.get():
            messagebox.showerror("Error", "Select folder and Excel first.")
            return
        
        self.log("🔍 Manual Extract: scanning folder...")
        self.status_label.config(text="🟡 Processing...")
        self.update_idletasks()
        
        rows = []
        count = 0
        errors = 0
        
        for fname in os.listdir(self.folder.get()):
            path = os.path.join(self.folder.get(), fname)
            if os.path.isfile(path):
                try:
                    row = self._process_file_to_row(path)
                    rows.append(row)
                    count += 1
                    self.log(f"✓ Processed: {os.path.basename(path)}")
                except Exception as e:
                    errors += 1
                    self.log(f"✗ Error: {os.path.basename(path)} -> {e}")
        
        ExcelWriter(self.excel.get()).write_all_sorted(rows)
        self.log(f"✓ Complete. Processed: {count} | Errors: {errors}")
        self.status_label.config(text="🟢 Ready")
        messagebox.showinfo("Success", f"Processed: {count} files\nErrors: {errors}")

    def refresh_sheet_sorted_from_folder(self):
        rows = []
        for fname in os.listdir(self.folder.get()):
            path = os.path.join(self.folder.get(), fname)
            if os.path.isfile(path):
                try: 
                    rows.append(self._process_file_to_row(path))
                except Exception as e: 
                    self.log(f"Error: {fname} -> {e}")
        ExcelWriter(self.excel.get()).write_all_sorted(rows)

    def rescan(self):
        if not self.folder.get() or not self.excel.get():
            messagebox.showerror("Error", "Select folder and Excel first.")
            return
        
        self.log("🔄 Rescanning folder and rebuilding Excel...")
        self.status_label.config(text="🟡 Scanning...")
        self.update_idletasks()
        
        self.refresh_sheet_sorted_from_folder()
        
        self.log("✓ Rescan complete.")
        self.status_label.config(text="🟢 Ready")

    def start_monitor(self):
        if not self.folder.get() or not self.excel.get():
            messagebox.showerror("Error", "Select folder and Excel first.")
            return
        
        self.log("⏳ Initial extraction...")
        self.status_label.config(text="🟡 Extracting...")
        self.update_idletasks()
        
        self.refresh_sheet_sorted_from_folder()
        
        self.log("✓ Extraction complete. Starting monitor...")
        self.observer = Observer()
        self.observer.schedule(FolderHandler(self), self.folder.get(), recursive=False)
        self.observer.start()
        self.stop_flag.clear()
        self.is_monitoring = True
        self.btn_start.config(state="disabled")
        self.status_label.config(text="🟢 Monitoring Active")

        def loop():
            while not self.stop_flag.is_set():
                if self.files_queue:
                    for _ in range(len(self.files_queue)):
                        p = self.files_queue.pop(0)
                        if os.path.isfile(p):
                            try:
                                _ = self._process_file_to_row(p)
                                self.log(f"✓ Processed (queued): {os.path.basename(p)}")
                            except Exception as e:
                                self.log(f"✗ Error processing {os.path.basename(p)} -> {e}")
                    try:
                        self.refresh_sheet_sorted_from_folder()
                        self.log("📊 Excel updated (sorted).")
                    except Exception as e:
                        self.log(f"✗ Error updating Excel: {e}")
                time.sleep(1.0)

        if (self.loop_thread is None) or (not self.loop_thread.is_alive()):
            self.loop_thread = threading.Thread(target=loop, daemon=True)
            self.loop_thread.start()

    def stop_monitor(self):
        if self.observer:
            self.observer.stop()
            self.observer = None
        self.stop_flag.set()
        self.is_monitoring = False
        self.btn_start.config(state="normal")
        self.log("⏹ Monitoring stopped.")
        self.status_label.config(text="🟢 Ready")

    def pick_folder(self):
        d = filedialog.askdirectory()
        if d: 
            self.folder.set(d)
            self.log(f"✓ Folder: {d}")

    def pick_excel(self):
        f = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if f: 
            self.excel.set(f)
            self.log(f"✓ Excel: {f}")

if __name__ == "__main__":
    app = App()
    app.mainloop()
