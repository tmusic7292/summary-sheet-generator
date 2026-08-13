import os
os.environ["NUMBA_DISABLE_JIT"] = "1"

#!/usr/bin/env python3
"""
Summary Sheet Tool v11.0 — Created by Pawan
Features:
- Real-time extraction via watchdog
- Remark = filename
- No skipping; all files processed
- Excel auto-extract: searches sheet for headers matching ADDITIONAL_HEADERS, value at row+2
- TXT auto-extract: opens in Excel-like row/col, searches headers, extracts value (like Excel)
- Enhanced PDF extraction: Emission Limits + Specific Results + Mass Results + Fuel Consumption
- Sorted final Excel by Test ID
- GUI safe (no main-thread errors)
- Modern dark-themed UI with professional styling
- Excel open detection (waits and shows message)
"""

import os
import re
import time
import threading
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pdfplumber
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# --- CONFIG ---
APP_TITLE = "Summary Sheet Tool v11.0 — Created by Pawan"
ADDITIONAL_HEADERS = [
    "Injector 1", "Injector 2", "Injector 3", "Injector 4",
    "DOC", "DPF", "Engine Number", "Dataset"
]
HEADERS = ["S No.", "Date", "Test ID", "Test Name", "NOx", "THC", "CO", "PM", "PN", "Remark"] + ADDITIONAL_HEADERS

# Enhanced PDF headers for additional extraction
PDF_ADDITIONAL_HEADERS = [
    "CO2", "COL", "NO2", "NO", "O2", "HC",  # From specific results
    "CH4", "Mass_NOx", "Mass_PM", "Mass_HC",  # From mass results
    "Fuel_Consumption", "Brake_Specific_FC"  # From fuel consumption
]

# --- COLOR THEME (Modern Dark) ---
THEME = {
    "bg_dark": "#1e1e1e",
    "bg_lighter": "#2d2d2d",
    "bg_input": "#3d3d3d",
    "fg_text": "#e0e0e0",
    "fg_accent": "#00d4ff",
    "fg_success": "#00ff88",
    "fg_error": "#ff4466",
    "btn_bg": "#0066cc",
    "btn_hover": "#0080ff",
}

# ----------- PDF PARSING - ENHANCED --------
def parse_pdf_comprehensive(path: str):
    """
    Extract data from all sections:
    1. Emission Limits (Stage V NRE)
    2. Specific Results
    3. Mass Results
    4. Fuel Consumption
    """
    emission_vals = {}
    specific_vals = {h: "" for h in PDF_ADDITIONAL_HEADERS}
    remark_bits = []
    
    try:
        with pdfplumber.open(path) as pdf:
            for page_idx, page in enumerate(pdf.pages):
                txt = page.extract_text() or ""
                
                # ========== SECTION 1: Emission Limits ==========
                if "Emission limits - Stage V NRE" in txt:
                    lines = txt.splitlines()
                    for l in lines:
                        nums = re.findall(r"([0-9][0-9Ee\+\-\.]*\.?[0-9]*)", l)
                        if nums:
                            nums = [n for n in nums if n and n not in ['.', '']]
                        
                        if "CO uncorrected" in l or "CO " in l:
                            emission_vals["CO"] = nums[0] if nums else ""
                        elif "NOX" in l or "NOx" in l:
                            emission_vals["NOx"] = nums[0] if nums else ""
                        elif "Particulate mass" in l:
                            emission_vals["PM"] = nums[0] if nums else ""
                        elif "Particulate number" in l:
                            emission_vals["PN"] = nums[0] if nums else ""
                        elif "HC" in l and "THC" not in l:
                            emission_vals["HC"] = nums[0] if nums else ""
                        
                        if ("Passed" in l) or ("Failed" in l):
                            remark_bits.append(l.strip())
                
                # ========== SECTION 2: Specific Results ==========
                if "Specific results" in txt.lower():
                    lines = txt.splitlines()
                    for i, l in enumerate(lines):
                        nums = re.findall(r"([0-9]+\.?[0-9]*)", l)
                        
                        # Map specific result values
                        if "CH4" in l and nums:
                            specific_vals["CH4"] = nums[0]
                        elif "CO2" in l and "COL" not in l and nums:
                            specific_vals["CO2"] = nums[0]
                        elif "COL" in l and nums:
                            specific_vals["COL"] = nums[0]
                        elif "NO2" in l and nums:
                            specific_vals["NO2"] = nums[0]
                        elif "NO " in l and nums:
                            specific_vals["NO"] = nums[0]
                        elif "O2" in l and nums:
                            specific_vals["O2"] = nums[0]
                        elif re.search(r'\bHC\b', l) and "THC" not in l and nums:
                            specific_vals["HC"] = nums[0]
                
                # ========== SECTION 3: Mass Results ==========
                if "mass results" in txt.lower():
                    lines = txt.splitlines()
                    for l in lines:
                        nums = re.findall(r"([0-9]+\.?[0-9]*)", l)
                        
                        if "NOx" in l and nums:
                            specific_vals["Mass_NOx"] = nums[0]
                        elif "PM" in l and nums:
                            specific_vals["Mass_PM"] = nums[0]
                        elif "HC" in l and nums:
                            specific_vals["Mass_HC"] = nums[0]
                
                # ========== SECTION 4: Fuel Consumption ==========
                if "fuel consumption" in txt.lower():
                    lines = txt.splitlines()
                    for l in lines:
                        nums = re.findall(r"([0-9]+\.?[0-9]*)", l)
                        
                        if ("fuel consumption" in l.lower() and "mass" in l.lower()) and nums:
                            specific_vals["Fuel_Consumption"] = nums[0]
                        elif ("brake" in l.lower() and "specific" in l.lower()) and nums:
                            specific_vals["Brake_Specific_FC"] = nums[0]
    
    except Exception as e:
        return emission_vals.get("NOx", ""), "", emission_vals.get("CO", ""), \
               emission_vals.get("PM", ""), emission_vals.get("PN", ""), \
               f"PDF parse error: {e}", specific_vals
    
    remark = "; ".join(remark_bits) if remark_bits else "Header not found"
    return emission_vals.get("NOx", ""), "", emission_vals.get("CO", ""), \
           emission_vals.get("PM", ""), emission_vals.get("PN", ""), \
           remark, specific_vals

# ----------- HELPERS --------
def parse_filename(fname: str):
    stem = Path(fname).stem
    parts = stem.split("_")
    test_id = "_".join(parts[:3]) if len(parts) >= 3 else stem
    test_name = "_".join(parts[3:]) if len(parts) > 3 else ""
    try:
        date_val = datetime.fromtimestamp(Path(fname).stat().st_mtime).strftime("%d/%m/%Y")
    except Exception:
        date_val = ""
    return test_id, test_name, date_val

def natural_key(s: str):
    s = str(s)
    parts = re.split(r'(\d+)', s)
    return [int(p) if p.isdigit() else p.lower() for p in parts]

def normalize_header(text: str):
    if text is None:
        return ""
    return re.sub(r'[\s_\-]+', '', str(text).strip().lower())

def parse_excel_for_additional(path: str):
    results = {h: "" for h in ADDITIONAL_HEADERS}
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        normalized_map = {normalize_header(h): h for h in ADDITIONAL_HEADERS}
        remaining = set(normalized_map.keys())
        max_row, max_col = ws.max_row or 0, ws.max_column or 0
        for r in range(1, max_row + 1):
            if not remaining:
                break
            for c in range(1, max_col + 1):
                if not remaining:
                    break
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    continue
                norm = normalize_header(cell.value)
                if norm in remaining:
                    header_name = normalized_map[norm]
                    target_row = r + 2
                    try:
                        val = ws.cell(row=target_row, column=c).value
                        results[header_name] = "" if val is None else str(val).strip()
                    except Exception:
                        results[header_name] = ""
                    remaining.remove(norm)
        return results
    except (InvalidFileException, Exception):
        return results

def parse_txt_for_additional(path: str):
    results = {h: "" for h in ADDITIONAL_HEADERS}
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            lines = [l.strip() for l in f.readlines() if l.strip()]
        normalized_map = {normalize_header(h): h for h in ADDITIONAL_HEADERS}
        remaining = set(normalized_map.keys())
        for i, line in enumerate(lines):
            if not remaining:
                break
            parts = re.split(r'\t|,', line)
            for j, cell in enumerate(parts):
                norm = normalize_header(cell)
                if norm in remaining:
                    header_name = normalized_map[norm]
                    target_index = i + 2
                    if target_index < len(lines):
                        val_parts = re.split(r'\t|,', lines[target_index])
                        if j < len(val_parts):
                            results[header_name] = val_parts[j].strip()
                    remaining.remove(norm)
        return results
    except Exception:
        return results

# ----------- EXCEL WRITER --------
class ExcelWriter:
    def __init__(self, path: str):
        self.path = Path(path)
        all_headers = HEADERS + PDF_ADDITIONAL_HEADERS
        if not self.path.exists():
            wb = Workbook()
            ws = wb.active
            ws.append(all_headers)
            wb.save(self.path)

    def write_all_sorted(self, rows):
        rows.sort(key=lambda r: natural_key(r[1]))
        all_headers = HEADERS + PDF_ADDITIONAL_HEADERS
        wb = Workbook()
        ws = wb.active
        ws.append(all_headers)
        for i, row in enumerate(rows, start=1):
            ws.append([i] + row)
        wb.save(self.path)

# ----------- WATCHDOG --------
class FolderHandler(FileSystemEventHandler):
    def __init__(self, app): 
        self.app = app
    def on_created(self, e):
        if not e.is_directory:
            self.app.queue_file(e.src_path)
    def on_modified(self, e):
        if not e.is_directory:
            self.app.queue_file(e.src_path)

# ----------- MODERN UI APP --------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1100x700")
        self.resizable(True, True)
        self.configure(bg=THEME["bg_dark"])
        
        # Configure styles
        self._setup_styles()
        
        self.folder = tk.StringVar()
        self.excel = tk.StringVar()
        self.files_queue = []
        self.observer = None
        self.loop_thread = None
        self.stop_flag = threading.Event()
        self.is_monitoring = False
        self._build_ui()

    def _setup_styles(self):
        """Setup ttk styles for modern look"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure colors
        style.configure('TFrame', background=THEME["bg_dark"])
        style.configure('TLabel', background=THEME["bg_dark"], foreground=THEME["fg_text"])
        style.configure('TEntry', fieldbackground=THEME["bg_input"], foreground=THEME["fg_text"])
        style.configure('TButton', background=THEME["btn_bg"], foreground=THEME["fg_text"])
        style.map('TButton', background=[('active', THEME["btn_hover"])])
        style.configure('Header.TLabel', font=('Segoe UI', 11, 'bold'), foreground=THEME["fg_accent"])
        style.configure('Title.TLabel', font=('Segoe UI', 14, 'bold'), foreground=THEME["fg_accent"])

    def _build_ui(self):
        """Build modern dark-themed UI"""
        # Main container
        main_frame = ttk.Frame(self)
        main_frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Title
        title_label = ttk.Label(main_frame, text="📊 Summary Sheet Tool v11.0", style='Title.TLabel')
        title_label.pack(anchor="w", pady=(0, 20))
        
        # --- Input Section ---
        input_frame = ttk.LabelFrame(main_frame, text="Configuration", padding=15)
        input_frame.pack(fill="x", pady=(0, 15))
        
        # Folder selection
        folder_frame = ttk.Frame(input_frame)
        folder_frame.pack(fill="x", pady=8)
        ttk.Label(folder_frame, text="📁 Reports Folder:", style='Header.TLabel').pack(side="left", padx=(0, 10))
        folder_entry = ttk.Entry(folder_frame, textvariable=self.folder, width=70)
        folder_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Button(folder_frame, text="Browse", command=self.pick_folder, width=12).pack(side="left")
        
        # Excel selection
        excel_frame = ttk.Frame(input_frame)
        excel_frame.pack(fill="x", pady=8)
        ttk.Label(excel_frame, text="📄 Master Excel File:", style='Header.TLabel').pack(side="left", padx=(0, 10))
        excel_entry = ttk.Entry(excel_frame, textvariable=self.excel, width=70)
        excel_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Button(excel_frame, text="Browse", command=self.pick_excel, width=12).pack(side="left")
        
        # --- Control Buttons ---
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(0, 15))
        
        self.btn_start = ttk.Button(button_frame, text="▶ Start Monitor", command=self.start_monitor, width=20)
        self.btn_start.pack(side="left", padx=5)
        
        ttk.Button(button_frame, text="⏹ Stop", command=self.stop_monitor, width=20).pack(side="left", padx=5)
        ttk.Button(button_frame, text="🔄 Rescan Now", command=self.rescan, width=20).pack(side="left", padx=5)
        ttk.Button(button_frame, text="📥 Manual Extract", command=self.manual_extract, width=20).pack(side="left", padx=5)
        
        # --- Status Bar ---
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill="x", pady=(0, 10))
        self.status_label = ttk.Label(status_frame, text="⚫ Ready", foreground=THEME["fg_accent"])
        self.status_label.pack(side="left")
        
        # --- Log Output ---
        log_frame = ttk.LabelFrame(main_frame, text="Log Output", padding=10)
        log_frame.pack(fill="both", expand=True)
        
        self.txt = tk.Text(
            log_frame, 
            height=25, 
            width=130,
            bg=THEME["bg_input"],
            fg=THEME["fg_text"],
            insertbackground=THEME["fg_accent"],
            font=('Courier New', 9),
            relief="flat",
            borderwidth=0
        )
        self.txt.pack(fill="both", expand=True, pady=5)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.txt.yview)
        scrollbar.pack(side="right", fill="y")
        self.txt.config(yscrollcommand=scrollbar.set)

    def pick_folder(self):
        d = filedialog.askdirectory()
        if d: 
            self.folder.set(d)
            self.log(f"✓ Folder selected: {d}")

    def pick_excel(self):
        f = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if f: 
            self.excel.set(f)
            self.log(f"✓ Excel file selected: {f}")

    def log(self, msg):
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.txt.insert("end", f"[{timestamp}] {msg}\n")
        self.txt.see("end")
        self.update_idletasks()

    def queue_file(self, path):
        base = os.path.basename(path)
        if base.startswith("~$") or base.startswith("~") or base.endswith(".tmp"):
            return
        if path not in self.files_queue:
            self.files_queue.append(path)

    def _process_file_to_row(self, f):
        test_id, test_name, date = parse_filename(f)
        nox = thc = co = pm = pn = ""
        remark = os.path.basename(f)
        low = f.lower()
        pdf_additional = {h: "" for h in PDF_ADDITIONAL_HEADERS}
        add_values = {h: "" for h in ADDITIONAL_HEADERS}
        
        # Process PDF with enhanced extraction
        if low.endswith(".pdf"):
            nox, thc, co, pm, pn, _remark, pdf_additional = parse_pdf_comprehensive(f)
        
        # Process Excel
        if low.endswith((".xlsx", ".xlsm", ".xls")):
            add_values = parse_excel_for_additional(f)
        elif low.endswith(".txt"):
            add_values = parse_txt_for_additional(f)
        
        add_list = [add_values.get(h, "") for h in ADDITIONAL_HEADERS]
        pdf_list = [pdf_additional.get(h, "") for h in PDF_ADDITIONAL_HEADERS]
        
        return [date, test_id, test_name, nox, thc, co, pm, pn, remark] + add_list + pdf_list

    def manual_extract(self):
        if not self.folder.get() or not self.excel.get():
            messagebox.showerror("Error", "Select folder and Excel first.")
            return
        
        self.log("🔍 Manual Extract: scanning folder...")
        self.status_label.config(text="🟡 Processing...")
        self.update_idletasks()
        
        rows = []
        count = 0
        errors = 0
        
        for fname in os.listdir(self.folder.get()):
            path = os.path.join(self.folder.get(), fname)
            if os.path.isfile(path):
                try:
                    row = self._process_file_to_row(path)
                    rows.append(row)
                    count += 1
                    self.log(f"✓ Processed: {os.path.basename(path)}")
                except Exception as e:
                    errors += 1
                    self.log(f"✗ Error: {os.path.basename(path)} -> {e}")
        
        ExcelWriter(self.excel.get()).write_all_sorted(rows)
        self.log(f"✓ Manual Extract done. Processed: {count} | Errors: {errors}")
        self.status_label.config(text="🟢 Ready")
        messagebox.showinfo("Success", f"Manual Extract Complete\nProcessed: {count} files\nErrors: {errors}")

    def refresh_sheet_sorted_from_folder(self):
        rows = []
        for fname in os.listdir(self.folder.get()):
            path = os.path.join(self.folder.get(), fname)
            if os.path.isfile(path):
                try: 
                    rows.append(self._process_file_to_row(path))
                except Exception as e: 
                    self.log(f"Error: {fname} -> {e}")
        ExcelWriter(self.excel.get()).write_all_sorted(rows)

    def rescan(self):
        if not self.folder.get() or not self.excel.get():
            messagebox.showerror("Error", "Select folder and Excel first.")
            return
        
        self.log("🔄 Rescanning folder and rebuilding Excel...")
        self.status_label.config(text="🟡 Scanning...")
        self.update_idletasks()
        
        self.refresh_sheet_sorted_from_folder()
        
        self.log("✓ Rescan complete.")
        self.status_label.config(text="🟢 Ready")

    def start_monitor(self):
        if not self.folder.get() or not self.excel.get():
            messagebox.showerror("Error", "Select folder and Excel first.")
            return
        
        self.log("⏳ Initial extraction...")
        self.status_label.config(text="🟡 Extracting...")
        self.update_idletasks()
        
        self.refresh_sheet_sorted_from_folder()
        
        self.log("✓ Extraction complete. Starting monitor...")
        self.observer = Observer()
        self.observer.schedule(FolderHandler(self), self.folder.get(), recursive=False)
        self.observer.start()
        self.stop_flag.clear()
        self.is_monitoring = True
        self.btn_start.config(state="disabled")
        self.status_label.config(text="🟢 Monitoring Active")

        def loop():
            while not self.stop_flag.is_set():
                if self.files_queue:
                    for _ in range(len(self.files_queue)):
                        p = self.files_queue.pop(0)
                        if os.path.isfile(p):
                            try:
                                _ = self._process_file_to_row(p)
                                self.log(f"✓ Processed (queued): {os.path.basename(p)}")
                            except Exception as e:
                                self.log(f"✗ Error processing {os.path.basename(p)} -> {e}")
                    try:
                        self.refresh_sheet_sorted_from_folder()
                        self.log("📊 Excel updated (sorted).")
                    except Exception as e:
                        self.log(f"✗ Error updating Excel: {e}")
                time.sleep(1.0)

        if (self.loop_thread is None) or (not self.loop_thread.is_alive()):
            self.loop_thread = threading.Thread(target=loop, daemon=True)
            self.loop_thread.start()

    def stop_monitor(self):
        if self.observer:
            self.observer.stop()
            self.observer = None
        self.stop_flag.set()
        self.is_monitoring = False
        self.btn_start.config(state="normal")
        self.log("⏹ Monitoring stopped.")
        self.status_label.config(text="🟢 Ready")

if __name__ == "__main__":
    app = App()
    app.mainloop()
